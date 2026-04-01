import os
import warnings
from copy import copy
from datetime import datetime
from pathlib import Path
from decimal import Decimal
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from sqlalchemy.orm import Session

from models import InquirySupplier, InquiryTask, InquiryTaskItem, InquiryRequest, Quotation, Supplier, LinkStatus


BASE_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = BASE_DIR / "static" / "templates"
CONTRACT_DIR = BASE_DIR / "static" / "contracts"
FONT_DIR = BASE_DIR / "static" / "fonts"
SIMSUN_PATH = FONT_DIR / "SimSun.ttf"
SYSTEM_SIMSUN_PATH = Path("C:/Windows/Fonts/simsun.ttc")
os.makedirs("static/contracts", exist_ok=True)
DEFAULT_TEMPLATE_CELLS = {
    "supplier_name": "E3",
    "buyer_name": "E4",
    "contract_no": "V3",
    "project_no": "F6",
    "total_amount_upper": "H9",
    "total_qty": "P9",
    "total_amount": "AA9",
    "sup_address": "I27",
    "sup_legal_rep": "I28",
    "sup_agent": "I29",
    "sup_phone": "I30",
    "sup_bank_name": "I31",
    "sup_bank_account": "I32",
    "sup_tax_id": "I33",
    "sup_fax": "I34",
    "sup_postal_code": "I35",
}
TEMPLATE_DYNAMIC_RULES = {
    "supplier_name": ("供方", 0, 3),
    "buyer_name": ("需方", 0, 3),
    "contract_no": ("合同号", 0, 4),
    "project_no": ("项目号", 0, 4),
    "total_amount_upper": ("合计人民币金额(大写)", 0, 6),
    "total_qty": ("数量", 2, 0),
    "total_amount": ("价税合计", 2, 0),
}


def _resolve_template_path() -> Path:
    candidates = [
        TEMPLATE_DIR / "合同模版.xlsx",
        TEMPLATE_DIR / "合同模版.XLSX",
        BASE_DIR / "合同模版.xlsx",
        BASE_DIR / "合同模版.XLS",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("合同模板文件不存在，请将‘合同模版.xlsx’放入 static/templates 目录")


def _normalize_text(value) -> str:
    return str(value or "").replace("：", ":").replace("\n", "").replace(" ", "").strip()


def _find_label_cell_openpyxl(ws, label: str, max_row: int = 20, max_col: int = 40):
    target = _normalize_text(label)
    row_limit = min(ws.max_row, max_row)
    col_limit = min(ws.max_column, max_col)
    for row_idx in range(1, row_limit + 1):
        for col_idx in range(1, col_limit + 1):
            current = _normalize_text(ws.cell(row_idx, col_idx).value)
            if not current:
                continue
            if target in current:
                return row_idx, col_idx
    return None


def _resolve_template_cells_for_openpyxl(ws) -> dict:
    template_cells = dict(DEFAULT_TEMPLATE_CELLS)
    for key, (label, row_offset, col_offset) in TEMPLATE_DYNAMIC_RULES.items():
        found = _find_label_cell_openpyxl(ws, label)
        if not found:
            continue
        row_idx = found[0] + row_offset
        col_idx = found[1] + col_offset
        if row_idx < 1 or col_idx < 1:
            continue
        template_cells[key] = f"{get_column_letter(col_idx)}{row_idx}"
    return template_cells


def _resolve_template_cells_for_win32(sheet) -> dict:
    template_cells = dict(DEFAULT_TEMPLATE_CELLS)
    row_limit = min(getattr(sheet.UsedRange, "Rows", sheet.UsedRange).Count, 20)
    col_limit = min(getattr(sheet.UsedRange, "Columns", sheet.UsedRange).Count, 40)
    label_positions = {}
    for row_idx in range(1, row_limit + 1):
        for col_idx in range(1, col_limit + 1):
            value = _normalize_text(sheet.Cells(row_idx, col_idx).Value)
            if not value:
                continue
            for _, (label, _, _) in TEMPLATE_DYNAMIC_RULES.items():
                if label in label_positions:
                    continue
                if _normalize_text(label) in value:
                    label_positions[label] = (row_idx, col_idx)
    for key, (label, row_offset, col_offset) in TEMPLATE_DYNAMIC_RULES.items():
        found = label_positions.get(label)
        if not found:
            continue
        row_idx = found[0] + row_offset
        col_idx = found[1] + col_offset
        if row_idx < 1 or col_idx < 1:
            continue
        template_cells[key] = f"{get_column_letter(col_idx)}{row_idx}"
    return template_cells


def _register_pdf_font() -> str:
    font_candidates = [SIMSUN_PATH, SYSTEM_SIMSUN_PATH]
    for font_path in font_candidates:
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont("SimSun", str(font_path.resolve())))
                return "SimSun"
            except Exception:
                continue
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    return "STSong-Light"


def _import_win32_modules():
    try:
        import pythoncom
        from win32com import client as win32
        return pythoncom, win32
    except Exception:
        return None, None


def _normalize_template_for_openpyxl(template_path: Path) -> Path:
    normalized_path = CONTRACT_DIR / "_normalized_template.xlsx"
    pythoncom, win32 = _import_win32_modules()
    if not pythoncom or not win32:
        raise RuntimeError("模板文件无法被 openpyxl 正常读取，且未检测到 pywin32")

    excel = None
    workbook = None
    try:
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(template_path.resolve()))
        normalized_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.SaveAs(str(normalized_path.resolve()), FileFormat=51)
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()
    return normalized_path


def _safe_load_workbook(file_path: Path):
    try:
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb = load_workbook(file_path)
        has_invalid_spec_warning = any("invalid specification" in str(w.message).lower() for w in caught)
        if has_invalid_spec_warning:
            normalized_template = _normalize_template_for_openpyxl(file_path)
            return load_workbook(normalized_template)
        return wb
    except Exception:
        normalized_template = _normalize_template_for_openpyxl(file_path)
        return load_workbook(normalized_template)


def _resolve_deal_link(db: Session, inquiry_id: int) -> InquirySupplier:
    deal_link = db.query(InquirySupplier).filter(
        InquirySupplier.id == inquiry_id,
        InquirySupplier.status == LinkStatus.DEAL
    ).first()
    if deal_link:
        return deal_link

    deal_link = db.query(InquirySupplier).filter(
        InquirySupplier.task_id == inquiry_id,
        InquirySupplier.status == LinkStatus.DEAL
    ).order_by(InquirySupplier.id.desc()).first()
    if deal_link:
        return deal_link

    raise ValueError("未找到已成交的询价记录")


def _to_decimal(val) -> Decimal:
    if val is None:
        return Decimal("0")
    return Decimal(str(val))


def _format_delivery_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            pass
    text = str(value).strip()
    if not text:
        return ""
    return text[:10]


def _to_chinese_upper_amount(amount: Decimal) -> str:
    digits = "零壹贰叁肆伍陆柒捌玖"
    units = ["", "拾", "佰", "仟"]
    big_units = ["", "万", "亿", "兆"]
    integer = int(amount.quantize(Decimal("1")))
    if integer == 0:
        return "零圆整"
    groups = []
    while integer > 0:
        groups.append(integer % 10000)
        integer //= 10000
    text_parts = []
    for gi in range(len(groups) - 1, -1, -1):
        group = groups[gi]
        if group == 0:
            if text_parts and not text_parts[-1].endswith("零"):
                text_parts.append("零")
            continue
        group_text = ""
        zero_flag = False
        for pos in range(3, -1, -1):
            divisor = 10 ** pos
            n = group // divisor
            group %= divisor
            if n == 0:
                zero_flag = True
            else:
                if zero_flag and group_text and not group_text.endswith("零"):
                    group_text += "零"
                zero_flag = False
                group_text += digits[n] + units[pos]
        text_parts.append(group_text + big_units[gi])
    return f"{''.join(text_parts).rstrip('零')}圆整"


def _collect_contract_payload(db: Session, link: InquirySupplier) -> dict:
    task = db.query(InquiryTask).filter(InquiryTask.id == link.task_id).first()
    supplier = db.query(Supplier).filter(Supplier.id == link.supplier_id).first()
    if not task or not supplier:
        raise ValueError("询价任务或供应商信息不存在")

    quotes = db.query(Quotation).filter(
        Quotation.inquiry_supplier_id == link.id,
        Quotation.round == link.current_round
    ).all()
    if not quotes:
        quotes = db.query(Quotation).filter(
            Quotation.inquiry_supplier_id == link.id
        ).order_by(Quotation.round.desc(), Quotation.id.asc()).all()
        if quotes:
            max_round = quotes[0].round
            quotes = [q for q in quotes if q.round == max_round]
    if not quotes:
        raise ValueError("未找到该成交供应商的报价数据")

    project_no = ""
    project_name = ""
    buyer_name = "俊朗电气有限公司"
    items = []
    total_amount = Decimal("0")
    total_qty = Decimal("0")

    for idx, q in enumerate(quotes, start=1):
        task_item = db.query(InquiryTaskItem).filter(InquiryTaskItem.id == q.item_id).first()
        req = db.query(InquiryRequest).filter(InquiryRequest.id == task_item.request_id).first() if task_item else None
        material_name = req.material_name if req else ""
        material_code = req.material_code if req else ""
        qty = _to_decimal(req.qty if req and req.qty is not None else q.qty)
        price = _to_decimal(q.price)
        amount = qty * price
        total_qty += qty
        total_amount += amount

        if not project_no and req and req.project_info:
            project_no = str(req.project_info.get("number") or req.project_info.get("name") or "")
        if not project_name and req and req.project_info:
            project_name = str(req.project_info.get("name") or req.project_info.get("number") or "")

        items.append({
            "index": idx,
            "material_name": material_name,
            "material_code": material_code,
            "qty": float(qty),
            "price": float(price),
            "amount": float(amount),
            "delivery_date": _format_delivery_date(q.delivery_date or (req.delivery_date if req else None)),
        })

    return {
        "contract_no": f"HT-{link.task_id}-{link.id}-{datetime.now().strftime('%Y%m%d')}",
        "supplier_name": supplier.name,
        "buyer_name": buyer_name,
        "project_no": project_no,
        "project_name": project_name,
        "task_title": task.title,
        "items": items,
        "total_qty": float(total_qty),
        "total_amount": float(total_amount),
        "sup_address": link.address or "",
        "sup_legal_rep": link.legal_rep or "",
        "sup_agent": link.agent or "",
        "sup_phone": link.phone or "",
        "sup_bank_name": link.bank_name or "",
        "sup_bank_account": link.bank_account or "",
        "sup_tax_id": link.tax_id or "",
        "sup_fax": link.fax or "",
        "sup_postal_code": link.postal_code or "",
    }


def _fill_template_excel(payload: dict, output_xlsx: Path, template_path: Path = None) -> None:
    template_path = template_path or _resolve_template_path()
    wb = _safe_load_workbook(template_path)
    if not wb.worksheets:
        normalized_template = _normalize_template_for_openpyxl(template_path)
        wb = load_workbook(normalized_template)
    if not wb.worksheets:
        raise ValueError("模板文件不包含可写工作表")
    ws = wb.active if wb.active else wb.worksheets[0]
    template_cells = _resolve_template_cells_for_openpyxl(ws)

    def resolve_cell_ref(cell_ref: str) -> str:
        try:
            cell = ws[cell_ref]
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        return ws.cell(row=merged_range.min_row, column=merged_range.min_col).coordinate
            return cell_ref
        except Exception:
            return cell_ref

    def set_cell_value(cell_ref: str, value):
        try:
            target_ref = resolve_cell_ref(cell_ref)
            ws[target_ref] = value
        except Exception:
            return

    def set_item_cell_value(row_idx: int, col_idx: int, value):
        try:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                        return
            ws.cell(row=row_idx, column=col_idx).value = value
        except Exception:
            return

    def clone_row_style(source_row: int, target_row: int):
        try:
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
            for col in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=source_row, column=col)
                target_cell = ws.cell(row=target_row, column=col)
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
        except Exception:
            return

    set_cell_value(template_cells["supplier_name"], payload.get("supplier_name", ""))
    set_cell_value(template_cells["buyer_name"], payload.get("buyer_name", ""))
    set_cell_value(template_cells["contract_no"], payload.get("contract_no", ""))
    set_cell_value(template_cells["project_no"], payload.get("project_no") or payload.get("task_title", ""))
    items = payload.get("items", [])
    row = 42
    max_item_row = ws.max_row
    remark_row = None
    for r in range(42, ws.max_row + 1):
        marker = ws.cell(row=r, column=2).value
        if isinstance(marker, str) and "备注" in marker:
            remark_row = r
            max_item_row = r - 1
            break
    if items and remark_row is not None:
        current_capacity = max_item_row - row + 1
        if current_capacity < 0:
            current_capacity = 0
        remark_row_height = ws.row_dimensions[remark_row].height
        remark_row_values = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=remark_row, column=col).value
            if value not in (None, ""):
                remark_row_values[col] = value
        remark_row_merges = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == remark_row and merged_range.max_row == remark_row:
                remark_row_merges.append((merged_range.min_col, merged_range.max_col))
        base_row_merges = []
        if current_capacity > 0:
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row == row and merged_range.max_row == row:
                    base_row_merges.append((merged_range.min_col, merged_range.max_col))

        def reset_item_row_structure(target_row: int):
            clone_row_style(row, target_row)
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == target_row and merged_range.max_row == target_row:
                    ws.unmerge_cells(str(merged_range))
            for min_col, max_col in base_row_merges:
                ws.merge_cells(
                    start_row=target_row,
                    start_column=min_col,
                    end_row=target_row,
                    end_column=max_col
                )

        def reset_remark_row_structure(target_row: int):
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == target_row and merged_range.max_row == target_row:
                    ws.unmerge_cells(str(merged_range))
            ws.row_dimensions[target_row].height = remark_row_height
            for min_col, max_col in remark_row_merges:
                ws.merge_cells(
                    start_row=target_row,
                    start_column=min_col,
                    end_row=target_row,
                    end_column=max_col
                )
            for col, value in remark_row_values.items():
                set_item_cell_value(target_row, col, value)

        final_remark_row = remark_row
        if len(items) > current_capacity:
            extra_rows = len(items) - current_capacity
            ws.insert_rows(remark_row, amount=extra_rows)
            for insert_idx in range(extra_rows):
                target_row = remark_row + insert_idx
                reset_item_row_structure(target_row)
            max_item_row = remark_row + extra_rows - 1
            final_remark_row = remark_row + extra_rows
        reset_remark_row_structure(final_remark_row)
        for target_row in range(row + 1, row + len(items)):
            if target_row <= max_item_row:
                reset_item_row_structure(target_row)
    for item in items:
        if row > max_item_row:
            break
        set_item_cell_value(row, 2, item.get("index"))
        set_item_cell_value(row, 3, payload.get("project_no") or payload.get("task_title", ""))
        set_item_cell_value(row, 7, payload.get("project_name") or payload.get("task_title", ""))
        set_item_cell_value(row, 9, item.get("material_name", ""))
        set_item_cell_value(row, 11, item.get("material_code", ""))
        set_item_cell_value(row, 15, item.get("qty", 0))
        set_item_cell_value(row, 19, item.get("price", 0))
        set_item_cell_value(row, 26, item.get("amount", 0))
        set_item_cell_value(row, 29, item.get("delivery_date", ""))
        row += 1
    total_amount = _to_decimal(payload.get("total_amount", 0))
    total_qty = _to_decimal(payload.get("total_qty", 0))
    set_cell_value(template_cells["total_amount_upper"], _to_chinese_upper_amount(total_amount))
    set_cell_value(template_cells["total_qty"], float(total_qty))
    set_cell_value(template_cells["total_amount"], float(total_amount))
    set_cell_value(template_cells["sup_address"], payload.get("sup_address", ""))
    set_cell_value(template_cells["sup_legal_rep"], payload.get("sup_legal_rep", ""))
    set_cell_value(template_cells["sup_agent"], payload.get("sup_agent", ""))
    set_cell_value(template_cells["sup_phone"], payload.get("sup_phone", ""))
    set_cell_value(template_cells["sup_bank_name"], payload.get("sup_bank_name", ""))
    set_cell_value(template_cells["sup_bank_account"], payload.get("sup_bank_account", ""))
    set_cell_value(template_cells["sup_tax_id"], payload.get("sup_tax_id", ""))
    set_cell_value(template_cells["sup_fax"], payload.get("sup_fax", ""))
    set_cell_value(template_cells["sup_postal_code"], payload.get("sup_postal_code", ""))

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def _fill_template_excel_with_win32(payload: dict, output_xlsx: Path) -> bool:
    pythoncom, win32 = _import_win32_modules()
    if not pythoncom or not win32:
        return False
    template_path = _resolve_template_path()
    excel = None
    workbook = None
    try:
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(template_path.resolve()))
        sheet = workbook.Worksheets(1)
        template_cells = _resolve_template_cells_for_win32(sheet)

        def set_range_value(cell_ref: str, value):
            try:
                cell_range = sheet.Range(cell_ref)
                if bool(cell_range.MergeCells):
                    cell_range.MergeArea.Cells(1, 1).Value = value
                else:
                    cell_range.Value = value
            except Exception:
                return

        def set_item_cell_value(row_idx: int, col_idx: int, value):
            try:
                cell = sheet.Cells(row_idx, col_idx)
                if bool(cell.MergeCells):
                    cell.MergeArea.Cells(1, 1).Value = value
                else:
                    cell.Value = value
            except Exception:
                return

        set_range_value(template_cells["supplier_name"], payload.get("supplier_name", ""))
        set_range_value(template_cells["buyer_name"], payload.get("buyer_name", ""))
        set_range_value(template_cells["contract_no"], payload.get("contract_no", ""))
        set_range_value(template_cells["project_no"], payload.get("project_no") or payload.get("task_title", ""))
        items = payload.get("items", [])
        row = 42
        max_item_row = int(sheet.UsedRange.Rows.Count)
        remark_row = None
        for r in range(42, max_item_row + 1):
            marker = sheet.Cells(r, 2).Value
            if isinstance(marker, str) and "备注" in marker:
                remark_row = r
                max_item_row = r - 1
                break
        if items and remark_row is not None:
            current_capacity = max_item_row - row + 1
            if current_capacity < 0:
                current_capacity = 0
            if len(items) > current_capacity:
                extra_rows = len(items) - current_capacity
                sheet.Rows(f"{remark_row}:{remark_row + extra_rows - 1}").Insert()
                max_item_row = remark_row + extra_rows - 1
        for item in items:
            if row > max_item_row:
                break
            set_item_cell_value(row, 2, item.get("index"))
            set_item_cell_value(row, 3, payload.get("project_no") or payload.get("task_title", ""))
            set_item_cell_value(row, 7, payload.get("project_name") or payload.get("task_title", ""))
            set_item_cell_value(row, 9, item.get("material_name", ""))
            set_item_cell_value(row, 11, item.get("material_code", ""))
            set_item_cell_value(row, 15, item.get("qty", 0))
            set_item_cell_value(row, 19, item.get("price", 0))
            set_item_cell_value(row, 26, item.get("amount", 0))
            set_item_cell_value(row, 29, item.get("delivery_date", ""))
            row += 1
        total_amount = _to_decimal(payload.get("total_amount", 0))
        total_qty = _to_decimal(payload.get("total_qty", 0))
        set_range_value(template_cells["total_amount_upper"], _to_chinese_upper_amount(total_amount))
        set_range_value(template_cells["total_qty"], float(total_qty))
        set_range_value(template_cells["total_amount"], float(total_amount))
        set_range_value(template_cells["sup_address"], payload.get("sup_address", ""))
        set_range_value(template_cells["sup_legal_rep"], payload.get("sup_legal_rep", ""))
        set_range_value(template_cells["sup_agent"], payload.get("sup_agent", ""))
        set_range_value(template_cells["sup_phone"], payload.get("sup_phone", ""))
        set_range_value(template_cells["sup_bank_name"], payload.get("sup_bank_name", ""))
        set_range_value(template_cells["sup_bank_account"], payload.get("sup_bank_account", ""))
        set_range_value(template_cells["sup_tax_id"], payload.get("sup_tax_id", ""))
        set_range_value(template_cells["sup_fax"], payload.get("sup_fax", ""))
        set_range_value(template_cells["sup_postal_code"], payload.get("sup_postal_code", ""))
        output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        workbook.SaveAs(str(output_xlsx.resolve()), FileFormat=51)
        return True
    except Exception:
        return False
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        if pythoncom is not None:
            pythoncom.CoUninitialize()


def _fill_template_to_temp_excel(payload: dict) -> Path:
    temp_xlsx = CONTRACT_DIR / f"temp_filled_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:8]}.xlsx"
    template_path = _resolve_template_path()
    _fill_template_excel(payload, temp_xlsx, template_path=template_path)
    return temp_xlsx


def _export_excel_to_pdf_with_win32(xlsx_path: Path, output_pdf: Path) -> bool:
    pythoncom, win32 = _import_win32_modules()
    if not pythoncom or not win32:
        return False

    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    excel = None
    workbook = None
    try:
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(xlsx_path.resolve()))
        workbook.ExportAsFixedFormat(0, str(output_pdf.resolve()))
        return True
    except Exception:
        return False
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        if pythoncom is not None:
            pythoncom.CoUninitialize()


def _export_excel_to_pdf(xlsx_path: Path, output_pdf: Path) -> None:
    if _export_excel_to_pdf_with_win32(xlsx_path, output_pdf):
        return
    _render_pdf_with_reportlab(xlsx_path, output_pdf)


def _render_pdf_with_reportlab(xlsx_path: Path, output_pdf: Path) -> None:
    wb = _safe_load_workbook(xlsx_path)
    ws = wb.active if wb.active else wb.worksheets[0]
    template_cells = _resolve_template_cells_for_openpyxl(ws)
    font_name = _register_pdf_font()

    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(output_pdf.resolve()))
    c.setPageSize((595, 842))
    c.setFont(font_name, 11)

    start_x = 40
    y = 800
    base_line_spacing = 20

    c.setFont(font_name, 16)
    c.drawString(start_x + 210, y, "采购合同")
    y -= 32
    c.setFont(font_name, 11)

    def get_display_value_by_ref(cell_ref: str):
        try:
            cell = ws[cell_ref]
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return cell.value
        except Exception:
            return ""

    def get_display_value_by_pos(row_idx: int, col_idx: int):
        try:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return cell.value
        except Exception:
            return ""

    header_refs = [
        ("供方", template_cells["supplier_name"]),
        ("需方", template_cells["buyer_name"]),
        ("合同号", template_cells["contract_no"]),
        ("项目号", template_cells["project_no"]),
    ]
    for label, ref in header_refs:
        val = get_display_value_by_ref(ref) or ""
        c.setFont(font_name, 11)
        c.drawString(start_x, y, f"{label}：{val}")
        y -= base_line_spacing

    supplier_extra_refs = [
        ("供方地址", template_cells["sup_address"]),
        ("法定代表人", template_cells["sup_legal_rep"]),
        ("委托代理人", template_cells["sup_agent"]),
        ("联系电话", template_cells["sup_phone"]),
        ("开户银行", template_cells["sup_bank_name"]),
        ("账号", template_cells["sup_bank_account"]),
        ("税号", template_cells["sup_tax_id"]),
        ("传真", template_cells["sup_fax"]),
        ("邮编", template_cells["sup_postal_code"]),
    ]
    for label, ref in supplier_extra_refs:
        val = get_display_value_by_ref(ref) or ""
        c.setFont(font_name, 10)
        c.drawString(start_x, y, f"{label}：{val}")
        y -= 16

    y -= 10
    table_headers = ["序号", "项目号", "项目名称", "物料名称", "型号规格", "数量", "含税单价", "价税合计", "交货日期"]
    col_widths = [32, 62, 75, 95, 95, 48, 68, 70, 70]
    x = start_x
    for i, h in enumerate(table_headers):
        c.setFont(font_name, 11)
        c.drawString(x, y, str(h))
        x += col_widths[i]
    y -= base_line_spacing

    row = 42
    while row <= ws.max_row and y > 80:
        values = [
            get_display_value_by_pos(row, 2),
            get_display_value_by_pos(row, 3),
            get_display_value_by_pos(row, 7),
            get_display_value_by_pos(row, 9),
            get_display_value_by_pos(row, 11),
            get_display_value_by_pos(row, 15),
            get_display_value_by_pos(row, 19),
            get_display_value_by_pos(row, 26),
            get_display_value_by_pos(row, 29),
        ]
        if isinstance(values[0], str) and "备注" in values[0]:
            break
        if all(v in [None, ""] for v in values):
            row += 1
            continue
        x = start_x
        for i, v in enumerate(values):
            c.setFont(font_name, 10)
            c.drawString(x, y, "" if v is None else str(v))
            x += col_widths[i]
        row_height = ws.row_dimensions[row].height or 15
        line_spacing = max(base_line_spacing, int(row_height + 4))
        y -= line_spacing
        row += 1

    c.save()


async def generate_contract_pdf(db: Session, inquiry_id: int) -> str:
    link = _resolve_deal_link(db, inquiry_id)
    payload = _collect_contract_payload(db, link)

    output_pdf = CONTRACT_DIR / f"合同_{inquiry_id}.pdf"
    temp_xlsx = _fill_template_to_temp_excel(payload)
    try:
        _export_excel_to_pdf(temp_xlsx, output_pdf)
    finally:
        if temp_xlsx.exists():
            temp_xlsx.unlink()

    static_pdf_path = f"/static/contracts/{output_pdf.name}"
    link.contract_pdf = static_pdf_path
    if hasattr(link, "contract_pdf_path"):
        link.contract_pdf_path = static_pdf_path
    db.add(link)
    db.commit()
    db.refresh(link)
    return static_pdf_path


async def generate_contract_pdf_from_mock_data(mock_data: dict, output_filename: str = "test_result.pdf") -> str:
    payload = {
        "contract_no": mock_data.get("contract_no", f"TEST-{datetime.now().strftime('%Y%m%d%H%M%S')}"),
        "supplier_name": mock_data.get("supplier_name", ""),
        "buyer_name": mock_data.get("buyer_name", "需方"),
        "project_no": mock_data.get("project_no", ""),
        "task_title": mock_data.get("task_title", "测试合同"),
        "items": mock_data.get("items", []),
        "total_amount": float(mock_data.get("total_amount", 0)),
    }
    output_pdf = CONTRACT_DIR / output_filename
    temp_xlsx = _fill_template_to_temp_excel(payload)
    try:
        _export_excel_to_pdf(temp_xlsx, output_pdf)
    finally:
        if temp_xlsx.exists():
            temp_xlsx.unlink()
    return str(output_pdf)
